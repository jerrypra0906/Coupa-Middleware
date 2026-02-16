import sys
from pathlib import Path

try:
  import openpyxl
except ImportError:  # pragma: no cover
  import subprocess

  subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
  import openpyxl  # type: ignore


def main() -> int:
  wb_path = Path("docs") / "Supplier Item Staging.xlsx"
  if not wb_path.exists():
    print(f"FILE_NOT_FOUND: {wb_path}", file=sys.stderr)
    return 1

  wb = openpyxl.load_workbook(wb_path, data_only=True)
  print("SHEETS:", ", ".join(wb.sheetnames))

  ws = wb[wb.sheetnames[0]]
  for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    values = ["" if v is None else str(v) for v in row]
    print("\t".join(values))
    if i >= 80:
      break

  return 0


if __name__ == "__main__":
  raise SystemExit(main())


